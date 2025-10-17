from pathlib import Path
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.label import DataLabelList


# ============================================================
# FUNCIÓN INTERNA — LECTURA Y PARSEO DE XMLS SRI
# ============================================================
def _leer_xml(xml_path: Path) -> dict:
    """
    Lee un archivo XML del SRI (Recibidos o Emitidos) y extrae campos clave.
    Devuelve un diccionario con la información del comprobante.
    """
    d = {"archivo": xml_path.name}
    try:
        root = ET.parse(xml_path).getroot()
        ns = {"ns": root.tag.split("}")[0].strip("{")} if "}" in root.tag else {}

        # --- Campos generales ---
        d["fechaEmision"] = root.findtext(".//ns:fechaEmision", default="", namespaces=ns).strip()
        d["rucEmisor"] = root.findtext(".//ns:ruc", default="", namespaces=ns).strip()
        d["razonSocial"] = root.findtext(".//ns:razonSocial", default="", namespaces=ns).strip()
        d["rucReceptor"] = root.findtext(".//ns:identificacionComprador", default="", namespaces=ns).strip()
        d["razonSocialReceptor"] = root.findtext(".//ns:razonSocialComprador", default="", namespaces=ns).strip()
        d["claveAcceso"] = root.findtext(".//ns:claveAcceso", default="", namespaces=ns).strip()
        d["tipoDocumento"] = root.findtext(".//ns:codDoc", default="", namespaces=ns).strip()

        # --- Conversión segura a float ---
        def _num(val):
            try:
                return float(val.replace(",", "").strip()) if val else 0.0
            except Exception:
                return 0.0

        # --- Totales ---
        d["subtotal"] = _num(root.findtext(".//ns:totalSinImpuestos", default="0", namespaces=ns))
        d["total"] = _num(root.findtext(".//ns:importeTotal", default="0", namespaces=ns))
        d["iva"] = 0.0

        # --- Detección de IVA (código 2) ---
        for imp in root.findall(".//ns:totalImpuesto", namespaces=ns):
            codigo = imp.findtext("ns:codigo", default="", namespaces=ns)
            valor = imp.findtext("ns:valor", default="0", namespaces=ns)
            if codigo == "2":
                d["iva"] += _num(valor)

    except Exception as e:
        d["error"] = f"Error leyendo {xml_path.name}: {e}"
    return d


# ============================================================
# FUNCIÓN PRINCIPAL — CONSTRUCCIÓN DE REPORTE DESDE XML
# ============================================================
def construir_reporte(carpeta_mes: Path, excel_salida: Path):
    """
    Genera un Excel con:
      - Detalle de comprobantes (por cada XML)
      - Totales por emisor
      - Hoja de errores (si aplica)
      - Gráfico de barras con estilo SRI
    """
    rows = []
    for xml in carpeta_mes.rglob("*.xml"):
        rows.append(_leer_xml(xml))

    if not rows:
        print("⚠️ No se encontraron archivos XML en la carpeta.")
        return

    df = pd.DataFrame(rows)
    errores = df[df.get("error").notna()] if "error" in df.columns else pd.DataFrame()

    # --- Agrupar totales por emisor ---
    piv = (
        df.groupby(["rucEmisor", "razonSocial"], dropna=False)[["subtotal", "iva", "total"]]
        .sum()
        .reset_index()
    )

    # --- Guardar Excel con múltiples hojas ---
    with pd.ExcelWriter(excel_salida, engine="openpyxl") as xls:
        df.to_excel(xls, index=False, sheet_name="Detalle")
        piv.to_excel(xls, index=False, sheet_name="Totales por Emisor")
        if not errores.empty:
            errores.to_excel(xls, index=False, sheet_name="Errores")

    _ajustar_columnas_excel(excel_salida)
    _insertar_grafico_corporativo(excel_salida)
    print(f"✅ Reporte generado: {excel_salida.name}")


# ============================================================
# NUEVA FUNCIÓN — REPORTE DIRECTO PARA EMITIDOS (HTML → EXCEL)
# ============================================================
def construir_reporte_emitidos(df_emitidos: pd.DataFrame, excel_salida: Path):
    """
    Crea un reporte de Emitidos a partir de un DataFrame.
    Incluye totales por cliente y gráfico de resumen.
    """
    if df_emitidos.empty:
        print("⚠️ No hay datos de emitidos para generar reporte.")
        return

    # --- Limpieza de datos ---
    numeric_cols = ["Subtotal", "IVA", "Total"]
    for col in numeric_cols:
        if col in df_emitidos.columns:
            df_emitidos[col] = pd.to_numeric(df_emitidos[col], errors="coerce").fillna(0)

    # --- Totales por cliente ---
    piv = (
        df_emitidos.groupby("Cliente", dropna=False)[numeric_cols]
        .sum()
        .reset_index()
    )

    # --- Exportar a Excel ---
    with pd.ExcelWriter(excel_salida, engine="openpyxl") as xls:
        df_emitidos.to_excel(xls, index=False, sheet_name="Detalle Emitidos")
        piv.to_excel(xls, index=False, sheet_name="Totales por Cliente")

    _ajustar_columnas_excel(excel_salida)
    _insertar_grafico_corporativo(excel_salida)
    print(f"✅ Reporte Emitidos generado: {excel_salida.name}")


# ============================================================
# AJUSTE AUTOMÁTICO DE ANCHO DE COLUMNAS
# ============================================================
def _ajustar_columnas_excel(archivo_excel: Path):
    wb = load_workbook(archivo_excel)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max(10, min(max_length + 2, 50))
    wb.save(archivo_excel)


# ============================================================
# INSERCIÓN DE GRÁFICO CORPORATIVO SRI
# ============================================================
def _insertar_grafico_corporativo(archivo_excel: Path):
    wb = load_workbook(archivo_excel)
    hoja_grafico = None
    for nombre in ["Totales por Emisor", "Totales por Cliente"]:
        if nombre in wb.sheetnames:
            hoja_grafico = nombre
            break
    if not hoja_grafico:
        return

    ws = wb[hoja_grafico]
    num_filas = ws.max_row
    if num_filas < 2:
        wb.save(archivo_excel)
        return

    chart = BarChart()
    chart.title = "Totales por Emisor/Cliente (SRI Audit)"
    chart.x_axis.title = "Entidad"
    chart.y_axis.title = "Total ($)"
    chart.height = 10
    chart.width = 20

    data = Reference(ws, min_col=4, min_row=1, max_row=num_filas)
    cats = Reference(ws, min_col=2, min_row=2, max_row=num_filas)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # --- Estilo corporativo SRI ---
    azul_sri = "1E4AA8"
    gris_suave = "A0A0A0"
    chart.graphicalProperties = GraphicalProperties(ln=ColorChoice(prstClr=gris_suave))
    chart.graphicalProperties.solidFill = azul_sri
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, "H2")
    wb.save(archivo_excel)
